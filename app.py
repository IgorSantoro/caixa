"""
Painel Tributário — IRPJ e CSLL
================================

Como usar
---------
Coloque na mesma pasta deste arquivo os relatórios de pagamento:

    base1.xls  ->  1º trimestre de apuração
    base2.xls  ->  2º trimestre
    base3.xls  ->  3º trimestre
    base4.xls  ->  4º trimestre

Suba só os trimestres que já existirem; o painel se ajusta sozinho.
Também aceita .xlsx. O trimestre vem do ARQUIVO, não da data de pagamento —
a apuração do 1º tri é recolhida em abril, que já é o 2º tri do calendário.

Sobre o ícone de olho
---------------------
A tabela é HTML, e HTML dentro do Streamlit não consegue chamar função Python
diretamente. O olho é um link que recarrega a página com ?empresa=... — e ele
leva junto a seção, o período e os filtros ativos, para que nada se perca
quando o painel de detalhe abre.
"""

import html
import io
import os
import re
from urllib.parse import quote

import pandas as pd
import streamlit as st

# =========================================================================
# CONFIGURAÇÃO
# =========================================================================

# Empresas de cada consolidado, pelo CÓDIGO (o número que abre o cf_empresa
# no relatório). O nome vem da própria base, então basta o código aqui.
# ESA é montada automaticamente como a união de todos os grupos abaixo.
GRUPOS = {
    "Geração":      [126, 137, 161, 196, 197, 198, 199, 231, 232],
    "Soluções":     [15, 160],
    "Alsol":        [230, 233, 246, 254, 255, 256, 257, 258,
                     281, 282, 283, 284, 285, 286, 301, 302],
    "Nova Denerge": [176, 260],
    "EBIOGÁS":      [266, 267, 308],
    "EDISGAS":      [268, 269, 303, 304],
    "Sobradinho":   [126, 236, 237, 238, 239, 240],
    "EPNE":         [27, 274],
    "ETE":          [216, 217, 224, 225, 228, 241, 242, 243,
                     247, 261, 262, 263, 271, 272, 273],
    "GEMINI":       [248, 249, 250, 251, 252, 253],
    "REDE":         [168, 170, 178, 182, 184, 190, 191, 193],
}
GRUPOS["ESA"] = sorted({cod for lista in GRUPOS.values() for cod in lista})

# Códigos DARF considerados. Qualquer outro código é descartado.
IMPOSTO_POR_CODIGO = {"2362": "IRPJ", "2089": "IRPJ", "2484": "CSLL", "2372": "CSLL"}

# Colunas de valor, na ordem em que aparecem nas exportações.
VALORES = ["vlr_principal", "vlr_multa", "vlr_juro_encargo",
           "vlr_outra_entidade", "vlr_total"]

# Deixe None para inferir o exercício pelas datas de pagamento.
EXERCICIO = None

PASTA = os.path.dirname(os.path.abspath(__file__))
ARQUIVOS = {t: [f"base{t}.xls", f"base{t}.xlsx"] for t in (1, 2, 3, 4)}
NOME_TRI = {1: "1º trimestre", 2: "2º trimestre",
            3: "3º trimestre", 4: "4º trimestre"}
SECOES = ["Empresas", "Consolidados"]

ICONE_OLHO = (
    '<svg viewBox="0 0 24 24" width="17" height="17" fill="none" '
    'stroke="currentColor" stroke-width="1.7" stroke-linecap="round" '
    'stroke-linejoin="round">'
    '<path d="M1.7 12S5.3 5.8 12 5.8 22.3 12 22.3 12 18.7 18.2 12 18.2 1.7 12 1.7 12Z"/>'
    '<circle cx="12" cy="12" r="3.1"/></svg>'
)


# =========================================================================
# APARÊNCIA
# =========================================================================

st.set_page_config(layout="wide", page_title="Painel Tributário",
                   page_icon="📊", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;600&display=swap');

:root {
  --bg:#060810; --card:#0C0F1A; --line:#1A2035; --line-soft:#0F1424;
  --txt:#E2E8F4; --dim:#94A3B8; --mut:#4A5A72;
  --blue:#3A8FF5; --purple:#9B6EF3; --teal:#3ABFBF; --amber:#F0A429; --green:#34C77B;
}
*, html, body, [class*="css"] { font-family:'IBM Plex Sans', system-ui, sans-serif !important; }
.stApp { background:var(--bg); }
.block-container { padding:1.5rem 2rem 3rem !important; max-width:100% !important; }

/* ─────────── cabeçalho ─────────── */
.head { border-bottom:1px solid var(--line); padding-bottom:1rem; margin-bottom:1.4rem; }
.head h1 { font-size:1.15rem; font-weight:700; color:var(--txt); margin:0; letter-spacing:-.3px; }
.head .escopo { font-size:.85rem; color:var(--dim); margin-top:6px; }
.head .escopo b { color:var(--blue); font-weight:600; }

/* ─────────── cartões de topo ─────────── */
.cards { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr));
         gap:12px; margin-bottom:.6rem; }
.card { background:var(--card); border:1px solid var(--line); border-radius:10px;
        padding:1rem 1.2rem; border-top:2px solid var(--line); }
.card.b { border-top-color:var(--blue); } .card.g { border-top-color:var(--green); }
.card.p { border-top-color:var(--purple); } .card.t { border-top-color:var(--teal); }
.card .rot { font-size:.64rem; font-weight:700; color:var(--mut);
             text-transform:uppercase; letter-spacing:1px; }
.card .val { font-size:1.15rem; font-weight:700; color:var(--txt); margin-top:8px;
             font-family:'IBM Plex Mono',monospace; letter-spacing:-.5px; }
.card.b .val { color:var(--blue); } .card.g .val { color:var(--green); }
.card.p .val { color:var(--purple); } .card.t .val { color:var(--teal); }
.card .sub { font-size:.68rem; color:var(--mut); margin-top:5px; }

.encargos { font-size:.72rem; color:var(--mut); margin:0 0 1.2rem 2px; }
.encargos b { color:var(--amber); font-weight:600; font-family:'IBM Plex Mono',monospace; }

.sec { font-size:.66rem; font-weight:700; color:var(--mut); text-transform:uppercase;
       letter-spacing:1.2px; margin:1.4rem 0 .7rem; display:flex; align-items:center; gap:10px; }
.sec::after { content:""; flex:1; height:1px; background:var(--line-soft); }

/* ─────────── tabela ─────────── */
.wrap { background:#0A0D18; border:1px solid var(--line); border-radius:10px; overflow:hidden; }
table.t { width:100%; border-collapse:collapse; font-size:.86rem; }
table.t thead tr { background:#080B15; border-bottom:1px solid var(--line); }
table.t th { padding:10px 16px; text-align:left; font-size:.62rem; font-weight:700;
             color:var(--mut); text-transform:uppercase; letter-spacing:1px; }
table.t th.n, table.t td.n { text-align:right; }
table.t th.acao, table.t td.acao { text-align:center; width:74px; padding:10px 10px; }
table.t tbody tr { border-bottom:1px solid var(--line-soft); }
table.t tbody tr:hover { background:#0F1428; }
table.t td { padding:12px 16px; color:var(--dim); }
table.t td.nome { color:var(--txt); font-weight:600; }
table.t td.n { font-family:'IBM Plex Mono',monospace; font-size:.8rem; color:#7A8CA6; }
table.t td.tot { font-family:'IBM Plex Mono',monospace; font-size:.86rem;
                 font-weight:700; color:var(--blue); text-align:right; }
.share { font-size:.7rem; color:var(--mut); font-family:'IBM Plex Mono',monospace; }
.bar { height:2px; background:var(--line-soft); border-radius:2px; margin-top:7px; }
.bar > i { display:block; height:2px; border-radius:2px; background:var(--blue); }

/* ─────────── botão de olho ─────────── */
a.olho { display:inline-flex; align-items:center; justify-content:center;
         width:32px; height:32px; border-radius:8px; color:#33415A;
         border:1px solid transparent; text-decoration:none;
         transition:color .15s, background .15s, border-color .15s; }
table.t tbody tr:hover a.olho { color:var(--dim); }
a.olho:hover { color:var(--blue) !important; background:#0D2040; border-color:#1A3A6A; }
a.olho:focus-visible { outline:2px solid var(--blue); outline-offset:2px; }

.vazio { background:var(--card); border:1px dashed var(--line); border-radius:10px;
         padding:2.5rem 1.5rem; text-align:center; color:var(--mut); font-size:.88rem; }

/* ─────────── barra lateral ─────────── */
[data-testid="stSidebar"] { background:#080B14 !important; border-right:1px solid var(--line) !important; }
[data-testid="stSidebar"] * { color:var(--dim) !important; }

.marca { display:flex; align-items:center; gap:11px; padding:.2rem 0 1.1rem;
         border-bottom:1px solid var(--line); margin-bottom:.4rem; }
.marca .sigla { width:36px; height:36px; border-radius:9px; flex-shrink:0;
                background:linear-gradient(135deg,#14315C,#3A8FF5);
                display:flex; align-items:center; justify-content:center;
                font-size:.72rem; font-weight:700; color:#fff !important; letter-spacing:.5px; }
.marca .nome { font-size:.92rem; font-weight:700; color:var(--txt) !important; line-height:1.2; }
.marca .sub  { font-size:.68rem; color:var(--mut) !important; margin-top:2px; }

.rot-lateral { font-size:.62rem; font-weight:700; color:var(--mut) !important;
               text-transform:uppercase; letter-spacing:1.2px; margin:1.1rem 0 .5rem; }

/* período: o rádio da lateral vira uma lista de opções clicáveis */
[data-testid="stSidebar"] div[role="radiogroup"] { gap:5px !important; }
[data-testid="stSidebar"] div[role="radiogroup"] > label {
  background:var(--card); border:1px solid var(--line); border-radius:8px;
  padding:.5rem .8rem !important; margin:0 !important; width:100%;
  transition:border-color .15s, background .15s; }
[data-testid="stSidebar"] div[role="radiogroup"] > label:hover { border-color:#2A3555; }
[data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked) {
  background:#0D2040; border-color:#1A3A6A; }
[data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked) p {
  color:var(--blue) !important; font-weight:600 !important; }
[data-testid="stSidebar"] div[role="radiogroup"] p { font-size:.82rem !important; }
[data-testid="stSidebar"] div[role="radiogroup"] > label > div:first-child { display:none !important; }

/* resumo e chips da lateral */
.mini { background:var(--card); border:1px solid var(--line); border-radius:8px;
        padding:.65rem .85rem; margin-bottom:6px; }
.mini .r { font-size:.58rem; font-weight:700; color:var(--mut) !important;
           text-transform:uppercase; letter-spacing:1px; }
.mini .v { font-size:.95rem; font-weight:700; margin-top:4px;
           font-family:'IBM Plex Mono',monospace; }
.mini .v.b { color:var(--blue) !important; } .mini .v.t { color:var(--teal) !important; }

.chips { display:flex; gap:6px; }
.chip { flex:1; text-align:center; font-size:.68rem; font-weight:700; padding:6px 0;
        border-radius:7px; font-family:'IBM Plex Mono',monospace;
        background:var(--card); border:1px solid var(--line); color:#33415A !important; }
.chip.on { background:#0C2A1B; border-color:#1B4A32; color:var(--green) !important; }
.legenda { font-size:.66rem; color:var(--mut) !important; margin-top:8px; line-height:1.5; }
.aviso { font-size:.66rem; color:var(--amber) !important; margin-top:10px; line-height:1.5;
         border-left:2px solid #3A2A0D; padding-left:8px; }

/* ─────────── controles ─────────── */
.stSelectbox > div > div, .stTextInput > div > input, .stMultiSelect > div > div {
  background:var(--card) !important; border:1px solid #1E2640 !important;
  color:var(--txt) !important; border-radius:7px !important; }
.stTextInput > div > input:focus { border-color:var(--blue) !important; }
.stDownloadButton > button { background:#0D2040 !important; color:var(--blue) !important;
  border:1px solid #1A3A6A !important; border-radius:7px !important; font-size:.82rem !important; }
.stDownloadButton > button:hover { background:var(--blue) !important; color:#fff !important; }
.stButton > button { background:var(--card) !important; color:var(--dim) !important;
  border:1px solid #1E2640 !important; border-radius:7px !important; font-size:.82rem !important; }
.stButton > button:hover { border-color:var(--blue) !important; color:var(--blue) !important; }
[data-testid="stMetric"] { background:var(--card) !important; border:1px solid var(--line) !important;
  border-radius:9px !important; padding:.9rem 1.1rem !important; }
[data-testid="stMetricValue"] { color:var(--txt) !important; font-family:'IBM Plex Mono',monospace !important; }
[data-testid="stDataFrameResizable"] { border:1px solid var(--line) !important; border-radius:9px !important; }
[data-testid="stDialog"] div[role="dialog"] { background:var(--bg) !important;
  border:1px solid var(--line) !important; }

/* seções principais: rádio horizontal com cara de abas */
section[data-testid="stMain"] div[role="radiogroup"] {
  gap:3px !important; background:var(--card); border:1px solid var(--line);
  border-radius:9px; padding:3px; display:inline-flex !important; margin-bottom:.4rem; }
section[data-testid="stMain"] div[role="radiogroup"] > label {
  background:transparent; border-radius:6px; padding:.4rem 1.3rem !important; margin:0 !important; }
section[data-testid="stMain"] div[role="radiogroup"] > label:has(input:checked) { background:#1A2440; }
section[data-testid="stMain"] div[role="radiogroup"] > label:has(input:checked) p {
  color:var(--txt) !important; font-weight:600 !important; }
section[data-testid="stMain"] div[role="radiogroup"] p { font-size:.85rem !important; color:var(--mut) !important; }
section[data-testid="stMain"] div[role="radiogroup"] > label > div:first-child { display:none !important; }
</style>
""", unsafe_allow_html=True)


# =========================================================================
# FUNÇÕES DE APOIO
# =========================================================================

def moeda(v):
    """1234.5 -> '1.234,50' (padrão brasileiro)."""
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def moeda_col(v):
    """Formatação de coluna de dataframe: 1234.5 -> 'R$ 1.234,50'."""
    try:
        return f"R$ {moeda(float(v))}"
    except (TypeError, ValueError):
        return v


def esc(t):
    """Escapa texto antes de injetar em HTML."""
    return html.escape(str(t))


def codigo_de(nome):
    """'126 - PARQUE EOLICO...' -> 126. Retorna None se não houver número."""
    m = re.match(r"\s*0*(\d+)", str(nome or ""))
    return int(m.group(1)) if m else None


def posicao(lista, valor, padrao=0):
    """Índice seguro para os defaults dos seletores."""
    lista = list(lista)
    return lista.index(valor) if valor in lista else padrao


@st.cache_data(show_spinner=False)
def exportar_excel(dados: pd.DataFrame, aba: str = "Dados") -> bytes:
    """Gera o .xlsx com cabeçalho azul, colunas de valor em R$ e painel congelado."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dados.to_excel(writer, index=False, sheet_name=aba[:31])
        ws = writer.sheets[aba[:31]]
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 24

        for i, coluna in enumerate(dados.columns, start=1):
            letra = get_column_letter(i)
            cabecalho = ws.cell(row=1, column=i)
            cabecalho.fill = PatternFill("solid", fgColor="1F3A6A")
            cabecalho.font = Font(bold=True, color="FFFFFF", size=11)
            cabecalho.alignment = Alignment(horizontal="center", vertical="center")

            largura = len(str(coluna)) + 4
            if not dados.empty:
                largura = max(largura, int(dados[coluna].astype(str).str.len().max()) + 2)
            ws.column_dimensions[letra].width = min(largura, 42)

            if coluna in VALORES:
                for linha in range(2, len(dados) + 2):
                    ws.cell(row=linha, column=i).number_format = 'R$ #,##0.00'
    return buffer.getvalue()


def resumo(df):
    """Totais de um recorte qualquer, num dicionário."""
    total = df["vlr_total"].sum()
    return {
        "total": total,
        "irpj": df.loc[df["Imposto"] == "IRPJ", "vlr_total"].sum(),
        "csll": df.loc[df["Imposto"] == "CSLL", "vlr_total"].sum(),
        "multa": df["vlr_multa"].sum(),
        "juros": df["vlr_juro_encargo"].sum(),
        "empresas": df["cf_empresa"].nunique(),
    }


def agregar(df, coluna="cf_empresa"):
    """Uma linha por empresa (ou por outra coluna), já ordenada pelo total."""
    base = df.groupby(coluna)[["vlr_total", "vlr_multa", "vlr_juro_encargo"]].sum()
    por_imposto = df.pivot_table(index=coluna, columns="Imposto", values="vlr_total",
                                 aggfunc="sum", fill_value=0)
    base["IRPJ"] = por_imposto["IRPJ"] if "IRPJ" in por_imposto else 0.0
    base["CSLL"] = por_imposto["CSLL"] if "CSLL" in por_imposto else 0.0
    return base.sort_values("vlr_total", ascending=False)


# =========================================================================
# BLOCOS DE TELA
# =========================================================================

def mostrar_cartoes(r, rotulo_total="Total pago", encargos=True):
    """Cartões de topo. Multa e juros ficam na linha fina abaixo, porque na
    maioria dos períodos são zero e roubavam espaço dos totais."""
    pct_irpj = r["irpj"] / r["total"] * 100 if r["total"] else 0
    pct_csll = r["csll"] / r["total"] * 100 if r["total"] else 0

    st.markdown(f"""
    <div class="cards">
      <div class="card b"><div class="rot">{esc(rotulo_total)}</div>
        <div class="val">R$ {moeda(r['total'])}</div><div class="sub">IRPJ + CSLL</div></div>
      <div class="card g"><div class="rot">IRPJ</div>
        <div class="val">R$ {moeda(r['irpj'])}</div><div class="sub">{pct_irpj:.1f}% do total</div></div>
      <div class="card p"><div class="rot">CSLL</div>
        <div class="val">R$ {moeda(r['csll'])}</div><div class="sub">{pct_csll:.1f}% do total</div></div>
      <div class="card t"><div class="rot">Empresas</div>
        <div class="val">{r['empresas']}</div><div class="sub">com pagamento no período</div></div>
    </div>""", unsafe_allow_html=True)

    if not encargos:
        return
    if r["multa"] or r["juros"]:
        st.markdown(
            f'<div class="encargos">Acréscimos legais no período — '
            f'multa <b>R$ {moeda(r["multa"])}</b> · '
            f'juros e encargos <b>R$ {moeda(r["juros"])}</b></div>',
            unsafe_allow_html=True)
    else:
        st.markdown('<div class="encargos">Sem multa ou juros no período.</div>',
                    unsafe_allow_html=True)


def link_olho(empresa, contexto):
    """Monta o link da coluna Ações. Leva junto seção, período e filtros para
    que a tela volte exatamente como estava depois que o painel abrir."""
    parametros = {"empresa": str(empresa)}
    parametros.update({c: v for c, v in contexto.items() if v})
    consulta = "&".join(f"{c}={quote(str(v))}" for c, v in parametros.items())
    return (f'<a class="olho" href="?{consulta}" target="_self" '
            f'title="Ver detalhe de {esc(empresa)}" '
            f'aria-label="Ver detalhe de {esc(empresa)}">{ICONE_OLHO}</a>')


def mostrar_tabela(base, rotulo="Empresa", sufixos=None, contexto=None):
    """Tabela única usada em todas as listagens: empresas e consolidados.

    base     — saída de agregar()
    sufixos  — texto opcional ao lado do nome (ex.: '8 de 8 empresas')
    contexto — se informado, desenha a coluna Ações com o olho de detalhe
    """
    if base.empty:
        st.markdown('<div class="vazio">Nenhum pagamento encontrado.</div>',
                    unsafe_allow_html=True)
        return

    teto = base["vlr_total"].max()
    soma = base["vlr_total"].sum()
    linhas = ""
    for nome, l in base.iterrows():
        largura = l["vlr_total"] / teto * 100 if teto else 0
        parte = l["vlr_total"] / soma * 100 if soma else 0
        extra = f' <span class="share">· {esc(sufixos[nome])}</span>' if sufixos else ""
        acao = f'<td class="acao">{link_olho(nome, contexto)}</td>' if contexto else ""
        linhas += f"""
        <tr>
          <td class="nome">
            <div style="display:flex;justify-content:space-between;gap:10px;">
              <span>{esc(nome)}{extra}</span><span class="share">{parte:.1f}%</span>
            </div>
            <div class="bar"><i style="width:{largura:.1f}%"></i></div>
          </td>
          <td class="n">R$ {moeda(l['IRPJ'])}</td>
          <td class="n">R$ {moeda(l['CSLL'])}</td>
          <td class="n">R$ {moeda(l['vlr_multa'])}</td>
          <td class="n">R$ {moeda(l['vlr_juro_encargo'])}</td>
          <td class="tot">R$ {moeda(l['vlr_total'])}</td>
          {acao}
        </tr>"""

    coluna_acao = '<th class="acao">Ações</th>' if contexto else ""
    st.markdown(f"""
    <div class="wrap"><table class="t">
      <thead><tr>
        <th>{esc(rotulo)} <span style="opacity:.4;font-weight:400;">· participação</span></th>
        <th class="n">IRPJ</th><th class="n">CSLL</th>
        <th class="n">Multa</th><th class="n">Juros</th><th class="n">Total</th>
        {coluna_acao}
      </tr></thead>
      <tbody>{linhas}</tbody>
    </table></div>""", unsafe_allow_html=True)


def botao_exportar(df, nome_arquivo, chave, aba="Dados"):
    """Exporta o recorte com todas as colunas de valor, uma linha por empresa."""
    tabela = (df.groupby("cf_empresa")[VALORES].sum()
                .reset_index().sort_values("vlr_total", ascending=False))
    st.download_button(
        "⬇  Exportar (.xlsx)",
        data=exportar_excel(tabela, aba),
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=chave,
    )


def corpo_detalhe(empresa, dados, periodo_texto, sufixo, ano_completo):
    """Conteúdo do painel de detalhe da empresa."""
    st.markdown(
        f'<div style="font-size:.78rem;color:#4A5A72;margin:-.4rem 0 1rem;">'
        f'{esc(periodo_texto)}</div>', unsafe_allow_html=True)

    mostrar_cartoes(resumo(dados), "Total da empresa", encargos=False)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Principal", f"R$ {moeda(dados['vlr_principal'].sum())}")
    c2.metric("Multa", f"R$ {moeda(dados['vlr_multa'].sum())}")
    c3.metric("Juros e encargos", f"R$ {moeda(dados['vlr_juro_encargo'].sum())}")
    c4.metric("Outras entidades", f"R$ {moeda(dados['vlr_outra_entidade'].sum())}")

    if ano_completo and dados["Trimestre"].nunique() > 1:
        st.markdown('<div class="sec">Por trimestre</div>', unsafe_allow_html=True)
        por_tri = (dados.groupby("Trimestre")["vlr_total"].sum()
                        .rename(index=NOME_TRI).reset_index())
        por_tri.columns = ["Trimestre", "vlr_total"]
        st.dataframe(por_tri.style.format({"vlr_total": moeda_col}),
                     use_container_width=True, hide_index=True)

    st.markdown(f'<div class="sec">{len(dados)} lançamentos</div>', unsafe_allow_html=True)
    lancamentos = dados[["dat_pagamento", "cod_pagamento", "Imposto"] + VALORES].copy()
    lancamentos = lancamentos.sort_values("dat_pagamento")
    lancamentos["dat_pagamento"] = lancamentos["dat_pagamento"].dt.strftime("%d/%m/%Y")
    st.dataframe(lancamentos.style.format({v: moeda_col for v in VALORES}),
                 use_container_width=True, hide_index=True)

    st.download_button(
        "⬇  Exportar lançamentos (.xlsx)",
        data=exportar_excel(lancamentos, "Lançamentos"),
        file_name=f"lancamentos_{codigo_de(empresa) or 'empresa'}_{sufixo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="exp_detalhe")


def abrir_detalhe(empresa, dados, periodo_texto, sufixo, ano_completo):
    """Mostra o detalhe em janela sobreposta. A janela mudou de nome e de
    assinatura ao longo das versões do Streamlit, então tentamos as formas
    conhecidas e, se nenhuma servir, mostramos o mesmo conteúdo na página."""
    argumentos = (empresa, dados, periodo_texto, sufixo, ano_completo)
    janela = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)

    if callable(janela):
        for extras in ({"width": "large"}, {}):
            try:
                janela(str(empresa), **extras)(corpo_detalhe)(*argumentos)
                return
            except TypeError:
                continue
            except Exception:
                break

    st.markdown(f'<div class="sec">Detalhe · {esc(empresa)}</div>', unsafe_allow_html=True)
    corpo_detalhe(*argumentos)


# =========================================================================
# CARGA E TRATAMENTO
# =========================================================================

@st.cache_data(show_spinner=False)
def ler_planilha(caminho, _mtime):
    """_mtime entra na chave do cache: se o arquivo mudar, recarrega sozinho."""
    return pd.read_excel(caminho)


def caminho_de(nomes):
    for nome in nomes:
        p = os.path.join(PASTA, nome)
        if os.path.exists(p):
            return p
    return None


def carregar():
    """Lê base1..base4. Cai para base.xls (trimestre pela data) se não achar nenhum."""
    partes, trimestres = [], []
    for tri, nomes in ARQUIVOS.items():
        caminho = caminho_de(nomes)
        if caminho:
            parte = ler_planilha(caminho, os.path.getmtime(caminho)).copy()
            parte["Trimestre"] = tri
            partes.append(parte)
            trimestres.append(tri)

    if partes:
        return pd.concat(partes, ignore_index=True), trimestres, True

    legado = caminho_de(["base.xls", "base.xlsx"])
    if not legado:
        st.error("Nenhuma base encontrada. Coloque base1.xls (e os demais trimestres) "
                 "nesta mesma pasta.")
        st.stop()
    return ler_planilha(legado, os.path.getmtime(legado)).copy(), [], False


def tratar(df, por_arquivo):
    """Descarta cancelados, mantém só IRPJ/CSLL e normaliza tipos."""
    if "dsc_situacao" in df.columns:
        situacao = df["dsc_situacao"].astype(str).str.upper().str.strip()
        df = df[~situacao.str.startswith("CANCELAD")]

    for coluna in ["cod_pagamento", "dat_pagamento", "cf_empresa"] + VALORES:
        if coluna not in df.columns:
            df[coluna] = 0

    df = df.copy()
    df["cod_pagamento"] = (df["cod_pagamento"].astype(str).str.strip()
                             .str.replace(r"\.0$", "", regex=True))
    df = df[df["cod_pagamento"].isin(IMPOSTO_POR_CODIGO)].copy()
    df["Imposto"] = df["cod_pagamento"].map(IMPOSTO_POR_CODIGO)
    df["cod_empresa"] = df["cf_empresa"].map(codigo_de)

    for coluna in VALORES:
        df[coluna] = pd.to_numeric(df[coluna], errors="coerce").fillna(0)

    df["dat_pagamento"] = pd.to_datetime(df["dat_pagamento"], dayfirst=True, errors="coerce")
    sem_data = int(df["dat_pagamento"].isna().sum())
    df = df.dropna(subset=["dat_pagamento"])

    if not por_arquivo:
        df["Trimestre"] = df["dat_pagamento"].dt.quarter
    return df, sem_data


bruto, tris_carregados, POR_ARQUIVO = carregar()
df, sem_data = tratar(bruto, POR_ARQUIVO)

if df.empty:
    st.error("A base foi lida, mas nenhum pagamento de IRPJ ou CSLL sobrou após o "
             "tratamento. Confira os códigos de receita e a coluna de situação.")
    st.stop()

anos = sorted(df["dat_pagamento"].dt.year.unique())
exercicio = EXERCICIO or int(df["dat_pagamento"].dt.year.mode()[0])
sem_codigo = int(df["cod_empresa"].isna().sum())


# =========================================================================
# ESTADO VINDO DO LINK DO OLHO
# =========================================================================
# O olho recarrega a página com ?empresa=...&sec=...&p=...&q=...&imp=...&g=...
# Lemos tudo aqui, guardamos em variáveis e limpamos a URL — assim a janela
# não reabre sozinha nas próximas interações.

_url = dict(st.query_params)
if _url:
    st.query_params.clear()

pedido_detalhe = _url.get("empresa")
secao_inicial = _url.get("sec", SECOES[0])
periodo_inicial = _url.get("p")
busca_inicial = _url.get("q", "")
grupo_inicial = _url.get("g")
impostos_iniciais = [i for i in _url.get("imp", "IRPJ,CSLL").split(",") if i] or ["IRPJ", "CSLL"]


# =========================================================================
# BARRA LATERAL
# =========================================================================

with st.sidebar:
    st.markdown(
        '<div class="marca"><div class="sigla">IR</div>'
        '<div><div class="nome">Painel Tributário</div>'
        f'<div class="sub">IRPJ e CSLL · exercício {exercicio}</div></div></div>',
        unsafe_allow_html=True)

    if POR_ARQUIVO:
        opcoes = ["Ano completo"] + [NOME_TRI[t] for t in sorted(tris_carregados)]
    else:
        st.markdown('<div class="rot-lateral">Ano de pagamento</div>', unsafe_allow_html=True)
        ano_legado = st.selectbox("Ano de pagamento", anos, index=len(anos) - 1,
                                  label_visibility="collapsed")
        df = df[df["dat_pagamento"].dt.year == ano_legado]
        exercicio = ano_legado
        opcoes = ["Ano completo"] + [NOME_TRI[t] for t in
                                     sorted(df["Trimestre"].dropna().unique())]

    st.markdown('<div class="rot-lateral">Período</div>', unsafe_allow_html=True)
    periodo = st.radio("Período", opcoes, index=posicao(opcoes, periodo_inicial),
                       label_visibility="collapsed")

    if periodo == "Ano completo":
        recorte = df
        rotulo_periodo = f"todos os trimestres de {exercicio}"
        sufixo_arquivo = f"{exercicio}_ano"
    else:
        numero = next(n for n, texto in NOME_TRI.items() if texto == periodo)
        recorte = df[df["Trimestre"] == numero]
        rotulo_periodo = f"{periodo} de {exercicio}"
        sufixo_arquivo = f"{exercicio}_Q{numero}"

    r_periodo = resumo(recorte)
    st.markdown('<div class="rot-lateral">No período selecionado</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="mini"><div class="r">Total pago</div>'
        f'<div class="v b">R$ {moeda(r_periodo["total"])}</div></div>'
        f'<div class="mini"><div class="r">Empresas</div>'
        f'<div class="v t">{r_periodo["empresas"]}</div></div>',
        unsafe_allow_html=True)

    st.markdown('<div class="rot-lateral">Trimestres no sistema</div>', unsafe_allow_html=True)
    if POR_ARQUIVO:
        chips = "".join(
            f'<div class="chip{" on" if t in tris_carregados else ""}">Q{t}</div>'
            for t in (1, 2, 3, 4))
        st.markdown(f'<div class="chips">{chips}</div>', unsafe_allow_html=True)
        pendentes = [f"Q{t}" for t in (1, 2, 3, 4) if t not in tris_carregados]
        if pendentes:
            st.markdown(
                f'<div class="legenda">Falta enviar: {", ".join(pendentes)}. '
                f'Copie o arquivo para a pasta do painel e ele aparece aqui.</div>',
                unsafe_allow_html=True)
    else:
        st.markdown(
            '<div class="legenda">Lendo base.xls — o trimestre está sendo deduzido '
            'da data de pagamento.</div>', unsafe_allow_html=True)

    alertas = []
    if sem_data:
        alertas.append(f"{sem_data} lançamento(s) sem data válida foram ignorados.")
    if sem_codigo:
        alertas.append(f"{sem_codigo} lançamento(s) sem código de empresa ficam fora "
                       f"dos consolidados.")
    if alertas:
        st.markdown(f'<div class="aviso">{" ".join(alertas)}</div>', unsafe_allow_html=True)


# =========================================================================
# CABEÇALHO E TOTAIS
# =========================================================================

st.markdown(
    f'<div class="head"><h1>Pagamentos de IRPJ e CSLL</h1>'
    f'<div class="escopo">Mostrando <b>{esc(rotulo_periodo)}</b> — '
    f'{len(recorte)} lançamentos de {recorte["cf_empresa"].nunique()} empresas.</div></div>',
    unsafe_allow_html=True)

if recorte.empty:
    st.markdown('<div class="vazio">Nenhum pagamento neste período. '
                'Escolha outro período na barra lateral.</div>', unsafe_allow_html=True)
    st.stop()

mostrar_cartoes(r_periodo, "Total pago")

# Rádio no lugar de st.tabs: o olho recarrega a página, e só com estado
# controlável a seção aberta volta a ser a mesma depois que o painel fecha.
secao = st.radio("Seção", SECOES, index=posicao(SECOES, secao_inicial),
                 horizontal=True, label_visibility="collapsed")


# ─── EMPRESAS ────────────────────────────────────────────────────────────
if secao == "Empresas":
    col_busca, col_imposto, col_export = st.columns([3, 1.4, 1])
    with col_busca:
        busca = st.text_input("Procurar empresa", value=busca_inicial,
                              placeholder="Nome ou código", key="busca_empresas")
    with col_imposto:
        impostos = st.multiselect("Imposto", ["IRPJ", "CSLL"],
                                  default=impostos_iniciais, key="imposto_empresas")

    visao = recorte[recorte["Imposto"].isin(impostos)] if impostos else recorte

    with col_export:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        botao_exportar(visao, f"empresas_{sufixo_arquivo}.xlsx", "exp_empresas", "Empresas")

    tabela = agregar(visao)
    if busca.strip():
        alvo = busca.strip().upper()
        tabela = tabela[tabela.index.astype(str).str.upper().str.contains(alvo, regex=False)]

    st.markdown(f'<div class="sec">{len(tabela)} empresas · maior pagador primeiro · '
                f'o olho abre o detalhe</div>', unsafe_allow_html=True)
    mostrar_tabela(tabela, contexto={
        "sec": "Empresas", "p": periodo, "q": busca, "imp": ",".join(impostos)})


# ─── CONSOLIDADOS ────────────────────────────────────────────────────────
else:
    # Comparativo entre grupos. ESA fica fora da tabela por ser a soma de
    # todos os outros — apareceria como uma linha repetindo o total.
    comparativo, quantidade = [], {}
    for nome, codigos in GRUPOS.items():
        if nome == "ESA":
            continue
        parte = recorte[recorte["cod_empresa"].isin(codigos)]
        if parte.empty:
            continue
        agregado = agregar(parte).sum()
        agregado.name = nome
        comparativo.append(agregado)
        quantidade[nome] = f"{parte['cf_empresa'].nunique()} de {len(codigos)} empresas"

    st.markdown('<div class="sec">Cada consolidado no período</div>', unsafe_allow_html=True)
    if comparativo:
        mostrar_tabela(pd.DataFrame(comparativo).sort_values("vlr_total", ascending=False),
                       rotulo="Consolidado", sufixos=quantidade)
        st.caption("O código 126 pertence a Geração e a Sobradinho, então somar as linhas "
                   "conta essa empresa duas vezes — os cartões do topo contam uma só.")
    else:
        st.markdown('<div class="vazio">Nenhum consolidado teve pagamento neste período.</div>',
                    unsafe_allow_html=True)

    st.markdown('<div class="sec">Abrir um consolidado</div>', unsafe_allow_html=True)
    lista_grupos = list(GRUPOS)
    grupo = st.selectbox("Consolidado", lista_grupos,
                         index=posicao(lista_grupos, grupo_inicial),
                         key="grupo_escolhido", label_visibility="collapsed")
    dados_grupo = recorte[recorte["cod_empresa"].isin(GRUPOS[grupo])]

    if dados_grupo.empty:
        st.markdown(f'<div class="vazio">{esc(grupo)} não teve pagamento neste período.</div>',
                    unsafe_allow_html=True)
    else:
        mostrar_cartoes(resumo(dados_grupo), f"Total do {grupo}")
        presentes = set(dados_grupo["cod_empresa"].dropna())
        parados = [c for c in GRUPOS[grupo] if c not in presentes]
        if parados:
            st.caption(f"Sem movimento no período: código(s) "
                       f"{', '.join(str(c) for c in parados)}.")

        _, col_dl = st.columns([4, 1])
        with col_dl:
            botao_exportar(dados_grupo,
                           f"{grupo.lower().replace(' ', '_')}_{sufixo_arquivo}.xlsx",
                           "exp_grupo", "Consolidado")
        mostrar_tabela(agregar(dados_grupo), contexto={
            "sec": "Consolidados", "p": periodo, "g": grupo})


# ─── PAINEL DE DETALHE (disparado pelo olho) ─────────────────────────────
if pedido_detalhe:
    dados_empresa = recorte[recorte["cf_empresa"] == pedido_detalhe]
    if dados_empresa.empty:
        st.warning(f"{pedido_detalhe} não tem pagamento em {rotulo_periodo}.")
    else:
        abrir_detalhe(pedido_detalhe, dados_empresa,
                      f"Pagamentos em {rotulo_periodo}",
                      sufixo_arquivo, periodo == "Ano completo")
